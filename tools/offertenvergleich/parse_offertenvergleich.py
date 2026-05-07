"""
Parst Offertenvergleich_Burglen_2026.xlsx und erzeugt Marktpreis-Datenbank.

Quelle: D:/Offertenvergleich/__Vergleich/Offertenvergleich_Burglen_2026.xlsx
Output: Knowledge/sanierung/marktpreise_burglen_2026.json

Spaltenlayout (analysiert):
  C1=Pos, C2=Bezeichnung, C3=Einheit
  C4=Eigendevis-Menge, C5=Fretz-Menge, C6=GKS-Menge
  C7=Fretz-EP, C8=Fretz-Total
  C9=GKS-EP, C10=GKS-Total*, C11=GKS-Total**  (NPK-Variation)
  C12=iTS-Menge, C13=iTS-EP, C14=iTS-Total

Anbieter EPs werden als Marktpreise pro Position aggregiert (Min/Median/Max).
"""
import json
import statistics
import sys
from pathlib import Path

import openpyxl

XLSX = Path(r"D:/Offertenvergleich/__Vergleich/Offertenvergleich_Burglen_2026.xlsx")
OUT = Path(r"C:/Sewer-Studio_KI_4.2/Knowledge/sanierung/marktpreise_burglen_2026.json")


def is_number(v):
    if v is None:
        return False
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def num(v):
    return float(v) if is_number(v) else None


def parse_zone(ws, zone_name: str):
    """Liest alle Datenzeilen einer Zone-Sheet."""
    rows = []
    for r in range(7, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        # Subtotal-/Header-Zeilen ueberspringen
        if c1 is None and c2 is None:
            continue
        s1 = str(c1 or "").strip()
        if s1.endswith("Σ") or "/" in s1 and c2 is None:
            continue  # Block-Header oder Subtotal
        if not c2:  # ohne Bezeichnung -> kein Datum
            continue

        bez = str(c2).strip()
        einh = str(c3 or "").strip()

        eig_mng = num(ws.cell(row=r, column=4).value)
        fretz_mng = num(ws.cell(row=r, column=5).value)
        gks_mng = num(ws.cell(row=r, column=6).value)
        fretz_ep = num(ws.cell(row=r, column=7).value)
        fretz_tot = num(ws.cell(row=r, column=8).value)
        gks_ep = num(ws.cell(row=r, column=9).value)
        gks_tot = num(ws.cell(row=r, column=10).value)
        # Spalte 11 ist meist Total-Variante (Pauschale + Detailabrechnung)
        its_mng = num(ws.cell(row=r, column=12).value)
        its_ep = num(ws.cell(row=r, column=13).value)
        its_tot = num(ws.cell(row=r, column=14).value)

        rows.append({
            "zone": zone_name,
            "pos": s1 if s1 != "—" else None,
            "bezeichnung": bez,
            "einheit": einh,
            "eigendevis_menge": eig_mng,
            "anbieter": {
                "Fretz": {"menge": fretz_mng, "ep": fretz_ep, "total": fretz_tot},
                "GKS":   {"menge": gks_mng,   "ep": gks_ep,   "total": gks_tot},
                "iTS":   {"menge": its_mng,   "ep": its_ep,   "total": its_tot},
            },
        })
    return rows


def aggregate_market_prices(all_rows):
    """Pro Bezeichnung die EPs aller Anbieter sammeln und Min/Median/Max berechnen."""
    by_bez = {}
    for row in all_rows:
        key = (row["bezeichnung"].lower().strip(), row["einheit"].strip())
        if key not in by_bez:
            by_bez[key] = {
                "bezeichnung": row["bezeichnung"],
                "einheit": row["einheit"],
                "pos_codes": set(),
                "zonen": set(),
                "ep_samples": [],   # Liste von (anbieter, zone, ep)
                "eigendevis_mengen": [],
            }
        e = by_bez[key]
        if row["pos"]:
            e["pos_codes"].add(row["pos"])
        e["zonen"].add(row["zone"])
        if row["eigendevis_menge"] is not None:
            e["eigendevis_mengen"].append(row["eigendevis_menge"])
        for anbieter, vals in row["anbieter"].items():
            ep = vals["ep"]
            if ep is not None and ep > 0:
                e["ep_samples"].append({
                    "anbieter": anbieter,
                    "zone": row["zone"],
                    "ep": ep,
                })

    result = []
    for (bez_lower, einh), e in by_bez.items():
        eps = [s["ep"] for s in e["ep_samples"]]
        if not eps:
            continue
        result.append({
            "bezeichnung": e["bezeichnung"],
            "einheit": e["einheit"],
            "pos_codes": sorted(e["pos_codes"]),
            "zonen": sorted(e["zonen"]),
            "ep_samples": e["ep_samples"],
            "ep_min": min(eps),
            "ep_median": round(statistics.median(eps), 2),
            "ep_max": max(eps),
            "ep_anzahl": len(eps),
            "eigendevis_menge_typisch": (
                round(statistics.median(e["eigendevis_mengen"]), 2)
                if e["eigendevis_mengen"] else None
            ),
        })
    # Sortiere nach Pos-Codes
    result.sort(key=lambda x: (x["pos_codes"][0] if x["pos_codes"] else "z", x["bezeichnung"]))
    return result


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    if not XLSX.exists():
        print(f"FEHLER: Datei nicht gefunden: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    zone_sheets = [s for s in wb.sheetnames if s.startswith("Zone ")]
    print(f"Verarbeite {len(zone_sheets)} Zone-Sheets: {zone_sheets}")

    all_rows = []
    for sn in zone_sheets:
        rows = parse_zone(wb[sn], sn)
        all_rows.extend(rows)
        print(f"  {sn}: {len(rows)} Positionen")

    # Endsummen aus Uebersicht extrahieren
    uebersicht = wb["Übersicht"]
    endsummen = {}
    for r in range(7, 15):
        c1 = uebersicht.cell(row=r, column=1).value
        if c1 and "Zone" in str(c1) or (c1 and any(d in str(c1) for d in ["5.0", "5.1"])):
            endsummen[str(c1).strip()] = {
                "eigendevis": num(uebersicht.cell(row=r, column=2).value),
                "fretz_brutto": num(uebersicht.cell(row=r, column=3).value),
                "fretz_inkl_mwst": num(uebersicht.cell(row=r, column=4).value),
                "gks_brutto": num(uebersicht.cell(row=r, column=5).value),
                "gks_inkl_mwst": num(uebersicht.cell(row=r, column=6).value),
                "its_brutto": num(uebersicht.cell(row=r, column=7).value),
                "its_inkl_mwst": num(uebersicht.cell(row=r, column=8).value),
                "guenstigster": uebersicht.cell(row=r, column=9).value,
            }

    market_prices = aggregate_market_prices(all_rows)

    output = {
        "_meta": {
            "quelle": "Submission Bürglen 2026 - Sanierungsofferten",
            "auftraggeber": "Abwasser Uri",
            "anbieter": ["Fretz Kanal-Service", "GKS", "iTS"],
            "stand": "2026-04-28",
            "anzahl_positionen": len(market_prices),
            "anzahl_rohzeilen": len(all_rows),
            "hinweise": [
                "EP = Einheitspreis pro Anbieter; Mengen variieren wegen unterschiedlicher NPK-Strukturen.",
                "Aggregation: Pro Bezeichnung+Einheit Min/Median/Max der gueltigen Anbieter-EPs.",
                "Eigendevis-Mengen sind interne Schaetzung Abwasser Uri (vor Submission).",
            ],
        },
        "endsummen_pro_zone": endsummen,
        "marktpreise": market_prices,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nGeschrieben: {OUT}")
    print(f"  {len(market_prices)} Marktpreis-Eintraege")
    print(f"  {len(endsummen)} Zonen mit Endsummen")

    # Beispiel-Auszug
    print("\n=== Beispiel-Marktpreise (erste 8) ===")
    for mp in market_prices[:8]:
        print(f"  [{','.join(mp['pos_codes'])}] {mp['bezeichnung'][:50]:50} ({mp['einheit']}): "
              f"Min {mp['ep_min']:>7.2f} | Med {mp['ep_median']:>7.2f} | Max {mp['ep_max']:>7.2f} CHF "
              f"(n={mp['ep_anzahl']})")


if __name__ == "__main__":
    main()
