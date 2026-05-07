"""
Parst alle Auswertungs-Excels aus D:/Offertenvergleich/Zone_*/Auswertung_*/
und erzeugt ein KI-Trainings-Set mit historischen Sanierungen.

Output: Knowledge/sanierung/historische_sanierungen.json

Pro Haltung:
  - Identifikation: Haltungsname, Strasse, DN, Material, Nutzungsart, Laenge
  - Befund: Schadensbeschreibung, Zustandsklasse
  - Empfohlene Massnahme + Kosten
  - Mengen pro Massnahmen-Typ:
    Inliner_St, Inliner_m, Anschluesse_St, Manschetten, Kurzliner, Neubau_m

Verwendung: Lookup "Aehnliche historische Haltungen mit DN/Material/Schaden" -> Massnahme + Kosten.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT_DIR = Path(r"D:/Offertenvergleich")
OUT = Path(r"C:/Sewer-Studio_KI_4.2/Knowledge/sanierung/historische_sanierungen.json")


def num(v):
    if v is None:
        return None
    try:
        f = float(v)
        return f if f != 0 else 0
    except (TypeError, ValueError):
        return None


def parse_haltungen_sheet(ws, zone: str):
    """Parst das Haltungen-Sheet. Datenzeilen ab Zeile 12."""
    rows = []
    for r in range(12, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value  # NR
        c2 = ws.cell(row=r, column=2).value  # Haltungsname
        if c1 is None or c2 is None:
            continue
        try:
            int(c1)  # Filter: nur Zeilen mit echter Nummer
        except (TypeError, ValueError):
            continue

        rows.append({
            "zone": zone,
            "nr": int(c1),
            "haltung": str(c2).strip(),
            "strasse": str(ws.cell(row=r, column=3).value or "").strip(),
            "material": str(ws.cell(row=r, column=4).value or "").strip(),
            "dn_mm": num(ws.cell(row=r, column=5).value),
            "nutzungsart": str(ws.cell(row=r, column=6).value or "").strip(),
            "laenge_m": num(ws.cell(row=r, column=7).value),
            "inspektionsrichtung": str(ws.cell(row=r, column=8).value or "").strip(),
            "schaden_beschreibung": str(ws.cell(row=r, column=9).value or "").strip(),
            "zustandsklasse": num(ws.cell(row=r, column=10).value),
            "pruefungsresultat": str(ws.cell(row=r, column=11).value or "").strip(),
            "sanieren": str(ws.cell(row=r, column=12).value or "").strip(),
            "massnahme_text": str(ws.cell(row=r, column=13).value or "").strip(),
            "kosten_chf": num(ws.cell(row=r, column=14).value),
            "eigentuemer": str(ws.cell(row=r, column=15).value or "").strip(),
            "mengen": {
                "inliner_st": num(ws.cell(row=r, column=18).value),
                "inliner_m": num(ws.cell(row=r, column=19).value),
                "anschluesse_st": num(ws.cell(row=r, column=20).value),
                "manschetten": num(ws.cell(row=r, column=21).value),
                "kurzliner": num(ws.cell(row=r, column=22).value),
                "neubau_m": num(ws.cell(row=r, column=23).value),
            },
            "status": str(ws.cell(row=r, column=24).value or "").strip(),
        })
    return rows


def aggregate_by_profile(haltungen):
    """Aggregiert nach (DN-Klasse, Material, Nutzungsart): typische Massnahme + Kosten/m."""
    profiles = {}
    for h in haltungen:
        dn = h["dn_mm"] or 0
        dn_klasse = (
            "≤200" if dn <= 200 else
            "DN250-300" if dn <= 300 else
            "DN400+" if dn <= 500 else
            ">DN500"
        )
        key = (dn_klasse, h["material"] or "?", h["nutzungsart"] or "?")
        if key not in profiles:
            profiles[key] = {
                "dn_klasse": dn_klasse,
                "material": h["material"],
                "nutzungsart": h["nutzungsart"],
                "anzahl": 0,
                "haltungslaengen": [],
                "kosten_pro_m": [],
                "kosten_pro_haltung": [],
                "massnahmen_text": [],
                "anschluss_je_haltung": [],
            }
        p = profiles[key]
        p["anzahl"] += 1
        if h["laenge_m"]:
            p["haltungslaengen"].append(h["laenge_m"])
        if h["kosten_chf"] and h["kosten_chf"] > 0 and h["laenge_m"]:
            p["kosten_pro_m"].append(round(h["kosten_chf"] / h["laenge_m"], 2))
            p["kosten_pro_haltung"].append(h["kosten_chf"])
        if h["massnahme_text"]:
            p["massnahmen_text"].append(h["massnahme_text"])
        if h["mengen"]["anschluesse_st"]:
            p["anschluss_je_haltung"].append(h["mengen"]["anschluesse_st"])

    # Aggregate-Statistik
    import statistics
    result = []
    for (dn_kl, mat, nutz), p in sorted(profiles.items()):
        kosten_m = p["kosten_pro_m"]
        kosten_h = p["kosten_pro_haltung"]
        anschl = p["anschluss_je_haltung"]
        result.append({
            "dn_klasse": dn_kl,
            "material": mat,
            "nutzungsart": nutz,
            "anzahl_faelle": p["anzahl"],
            "haltungslaenge_median_m": round(statistics.median(p["haltungslaengen"]), 1) if p["haltungslaengen"] else None,
            "kosten_pro_m_median_chf": round(statistics.median(kosten_m), 2) if kosten_m else None,
            "kosten_pro_m_min_chf": min(kosten_m) if kosten_m else None,
            "kosten_pro_m_max_chf": max(kosten_m) if kosten_m else None,
            "kosten_pro_haltung_median_chf": round(statistics.median(kosten_h), 0) if kosten_h else None,
            "anschluss_je_haltung_median": round(statistics.median(anschl), 0) if anschl else None,
            "typische_massnahmen": list(set(p["massnahmen_text"]))[:3],
        })
    return result


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    auswertungs_files = list(ROOT_DIR.glob("Zone_*/Auswertung_*/*Auswertung*.xlsx"))
    auswertungs_files = [f for f in auswertungs_files if not f.name.startswith("~$")]
    print(f"Gefunden: {len(auswertungs_files)} Auswertungs-Dateien")

    all_haltungen = []
    for f in auswertungs_files:
        zone = re.search(r"Zone_(\d+\.\d+)", f.parent.parent.name)
        zone_str = f"Zone {zone.group(1)}" if zone else "Unbekannt"
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            if "Haltungen" not in wb.sheetnames:
                print(f"  WARN: {f.name} kein Haltungen-Sheet")
                continue
            rows = parse_haltungen_sheet(wb["Haltungen"], zone_str)
            all_haltungen.extend(rows)
            print(f"  {f.name}: {len(rows)} Haltungen aus {zone_str}")
        except Exception as ex:
            print(f"  FEHLER bei {f.name}: {ex}")

    profiles = aggregate_by_profile(all_haltungen)

    output = {
        "_meta": {
            "quelle": "Auswertungen Bürglen-Sanierungsplanung 2024-2026 (Abwasser Uri)",
            "stand": "2026-04-28",
            "anzahl_haltungen": len(all_haltungen),
            "anzahl_profile": len(profiles),
            "zonen": sorted(set(h["zone"] for h in all_haltungen)),
            "verwendung": "Aehnlichkeits-Lookup fuer KI-Sanierungsempfehlungen + Kostenkalibrierung.",
        },
        "haltungen": all_haltungen,
        "profile_aggregat": profiles,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nGeschrieben: {OUT}")
    print(f"  {len(all_haltungen)} historische Haltungen")
    print(f"  {len(profiles)} Profile (DN-Klasse × Material × Nutzungsart)")
    print(f"\n=== Top-Profile (Anzahl Faelle) ===")
    for p in sorted(profiles, key=lambda x: -x["anzahl_faelle"])[:8]:
        print(f"  {p['dn_klasse']:>10} | {p['material']:>5} | {p['nutzungsart']:>14} : "
              f"n={p['anzahl_faelle']:>3}, Median {p['kosten_pro_m_median_chf']} CHF/m, "
              f"Anschluesse {p['anschluss_je_haltung_median']}")


if __name__ == "__main__":
    main()
