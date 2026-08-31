# -*- coding: utf-8 -*-
"""Erzeugt die beiden Export-Vorlagen fuer Haltungen und Schaechte.

Warum ein Werkzeug und keine von Hand gepflegte Datei: In der alten Vorlage
liefen Farben und Formeln unbemerkt auseinander (Zustandsklasse 3 war bei den
Haltungen AEB135, bei den Schaechten A5A832), eine Zaehlformel reichte nur bis
Zeile 60 statt 500, zwei Eigentuemer-Zaehler waren feste Nullen und in einer
Summe stand ein kaputter Bezug. Solche Fehler sieht in einer .xlsx niemand.

Die Vorlage traegt das ganze Aussehen: Logo, Diagramme, Kennzahlenbloecke mit
Formeln, bedingte Formatierung, Titelband, Kopfzeile, Druckeinrichtung und
GENAU EINE gestaltete Musterzeile. Der C#-Export schreibt nur noch Werte und
kopiert den Stil der Musterzeile nach unten.

Aufruf:  python vorlage.py [--uebernehmen]
Ohne --uebernehmen landen die Dateien nur im Unterordner "ausgabe".
"""
import os
import shutil
import sys

import openpyxl
from openpyxl.utils import get_column_letter as SP

import stil as S
import werkzeug as W

HIER = os.path.dirname(os.path.abspath(__file__))
LOGO = os.path.join(HIER, "logo.png")
AUSGABE = os.path.join(HIER, "ausgabe")
ZIEL = os.path.abspath(os.path.join(HIER, "..", "..", "Export_Vorlage"))

# Bis hierhin reichen Zaehlformeln und bedingte Formatierung. Der Export
# begrenzt die Zeilenzahl ohnehin; lieber grosszuegig als zu knapp - genau
# daran ist die alte Vorlage gescheitert.
LETZTE_ZEILE = 5000

GRAU = "FFD6DCE4"
# Anzeigename und alle Schreibweisen, die dazu zaehlen. Der Anzeigename ist der
# amtliche Begriff des Kantons; die Kurzform steht daneben, weil sie in
# gewachsenen Projekten vorkommt. Gezaehlt und summiert wird ueber beide, damit
# weder ein nachgeschlagener noch ein alter Wert aus der Auswertung faellt.
EIGNER = [
    ("Abwasser Uri", ["Abwasser Uri", "AWU"]),
    ("Kanton Uri", ["Kanton Uri", "Kanton"]),
    ("Bund", ["Bund"]),
    ("Gemeinde", ["Gemeinde"]),
    ("Privat", ["Privat"]),
]
AUSFUEHRENDE = ["Abwasser Uri", "Kanalsanierer", "Baumeister", "Gartenbauer"]
BEDEUTUNG = {"0": "sofort", "1": "kurzfristig", "2": "mittelfristig",
             "3": "langfristig", "4": "kein Bedarf"}

# Die Ueberschriften sind lesbare Berichtsbegriffe. Die C#-Zuordnung erkennt
# die fachlichen Feldschluessel und historische Schreibweisen getrennt davon.
KOPF_HALTUNGEN = [
    "NR.", "Haltungsname (ID)", "Strasse", "Rohrmaterial", "DN mm", "Nutzungsart",
    "Haltungslänge m", "Inspektionsrichtung", "Primäre Schäden", "Zustandsklasse",
    "Prüfungsresultat", "Sanieren Ja/Nein", "Empfohlene Sanierungsmassnahmen",
    "Kosten", "Eigentümer", "Ausgeführt durch", "Bemerkungen", "Link",
    "Renovierung Inliner Stk.", "Renovierung Inliner m", "Anschlüsse verpressen",
    "Reparatur Manschette", "Linerendmanschette LEM", "Reparatur Kurzliner",
    "Erneuerung Neubau m", "offen/abgeschlossen", "Datum/Jahr",
]
# Spalte 10 ist "Zustandsklasse": bei Breite 9 brach die Ueberschrift mitten
# im Wort ("Zustandskla|sse").
BREITE_HALTUNGEN = [6, 18, 18, 16, 8, 17, 11, 17, 40, 12, 14, 9, 34, 13, 12, 16,
                    22, 11, 9, 9, 9, 9, 9, 9, 9, 14, 12]

KOPF_SCHAECHTE = [
    "NR.", "Funktion", "Schachtnummer", "Strasse", "Primäre Schäden",
    "Zustandsklasse", "Ja/Nein", "Massnahmen", "Kosten", "Eigentümer",
    "Ausgeführt durch", "Bemerkungen", "Link", "Abdeckung Stk.",
    "Belastungsklasse", "Status\noffen/abgeschlossen", "Ausführung\nDatum/Jahr",
]
BREITE_SCHAECHTE = [6, 16, 15, 18, 38, 12, 9, 32, 14, 12, 16, 24, 11, 12, 14, 15, 13]

# Feste Auswahl fuer die Verteilungsbloecke. Was nicht aufgefuehrt ist, faellt
# ueber die Restformel in "uebrige" - so stimmt die Summe immer.
MATERIALIEN = ["Zement", "Polyethylen", "Beton", "Polyvinylchlorid", "Polypropylen"]
FUNKTIONEN = ["Kontrollschacht", "Einstiegschacht"]

# Zwei fachlich belegte Wertefamilien bleiben unveraendert in den Daten stehen.
# Der Bericht fasst nur ihre gleichfarbigen Ampelstufen fuer die Kennzahl zusammen.
PRUEFUNGSGRUPPEN = [
    ("i.O. / Prüfung bestanden",
     ["i.O.", "Prüfung bestanden", "Pruefung bestanden"], "FF92D050", False),
    ("beobachten / Prüfung knapp nicht bestanden",
     ["beobachten", "Prüfung knapp nicht bestanden", "Pruefung knapp nicht bestanden"],
     "FFFFFF00", False),
    ("Sanierungsbedarf / Prüfung nicht bestanden",
     ["Sanierungsbedarf", "Prüfung nicht bestanden (grob undicht)",
      "Pruefung nicht bestanden (grob undicht)"], "FFFF0000", True),
    ("Keine", ["Keine"], "FFE7E6E6", False),
]

BALKEN_X, BALKEN_B, BALKEN_H = 8.4, 13.0, 1.55
MENGEN_X, MENGEN_B, MENGEN_H = 21.8, 10.0, 7.75
# Dritte Kachel rechts, damit die Kopfzeile die Blattbreite nutzt.
DRITT_X, DRITT_B = 32.1, 10.0
FORMAT_CHF_KURZ = '#,##0 "CHF";;""'


def _bereich(spalte, erste):
    return "$%s$%d:$%s$%d" % (SP(spalte), erste, SP(spalte), LETZTE_ZEILE)


def _zaehlt(bereich, wert, numerisch=False):
    """Zaehlt einen Wert.

    Der Vergleichswert steht IMMER in Anfuehrungszeichen - auch bei Zahlen.
    Excel behandelt ein Kriterium in Anfuehrungszeichen als Muster und trifft
    damit sowohl die Zahl 2 als auch den Text "2". Der Export schreibt die
    Zustandsklasse als Text, eine von Hand getippte Zahl waere eine echte Zahl:
    ohne diese Schreibweise zaehlt die Formel je nach Herkunft nur die Haelfte.
    """
    return '=COUNTIF(%s,"%s")' % (bereich, wert)


def _zaehlt_mehrere(bereich, werte):
    """Addiert Schreibweisen, ohne einen gespeicherten Zellwert umzudeuten."""
    return "=" + "+".join('COUNTIF(%s,"%s")' % (bereich, wert) for wert in werte)


def _summiert_mehrere(bereich, werte, summenbereich):
    """Addiert Schreibweisen, ohne einen gespeicherten Zellwert umzudeuten."""
    return "=" + "+".join(
        'SUMIF(%s,"%s",%s)' % (bereich, wert, summenbereich) for wert in werte)


def _erzwinge_neuberechnung(wb):
    """Kennzahlen und Diagrammquellen werden erst von Excel ausgerechnet."""
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True


def _rest(bereich, sichtbar):
    abzug = "".join('-COUNTIF(%s,"%s")' % (bereich, n) for n in sichtbar)
    return "=COUNTA(%s)%s" % (bereich, abzug)


def _verteilungszeilen(bereich, werte, farbtabelle, ersatz):
    zeilen, farben = [], []
    for i, name in enumerate(werte):
        farbe = farbtabelle.get(name, (None, False))[0] or ersatz[i % len(ersatz)]
        zeilen.append((name, farbe, S.weisse_schrift(farbe), _zaehlt(bereich, name)))
        farben.append(farbe)
    zeilen.append(("übrige", GRAU, False, _rest(bereich, werte)))
    farben.append(GRAU)
    return zeilen, farben


def _musterzeile(ws, spalten, rechts, mitte, umbruch, formate, link_spalte):
    """Eine einzige gestaltete Datenzeile. Der Export kopiert ihren Stil."""
    r = W.Z_DATEN
    ws.row_dimensions[r].height = 30
    for sp in range(1, spalten + 1):
        z = ws.cell(r, sp)
        # Feine Linien in beide Richtungen: bei 27 Spalten braucht das Auge
        # senkrechte Trenner, um in der Zeile zu bleiben. Sie sind heller als
        # die Zeilentrenner, damit kein Karogitter entsteht.
        z.border = S.ZEILENLINIE
        z.font = S.schrift(9)
        z.alignment = S.ausrichtung(
            "right" if sp in rechts else ("center" if sp in mitte else "left"),
            "top" if sp in umbruch else "center", umbruch=sp in umbruch)
        if sp in formate:
            z.number_format = formate[sp]
    verweis = ws.cell(r, link_spalte)
    verweis.font = S.schrift(9, farbe="FF0563C1", unterstrichen="single")
    verweis.alignment = S.ausrichtung("center")


def _farbregeln(ws, spalten_regeln, erste):
    for ziel, quelle, zuordnung, numerisch in spalten_regeln:
        bereich = "%s%d:%s%d" % (SP(ziel), erste, SP(ziel), LETZTE_ZEILE)
        W.farbregeln(ws, bereich, "$%s%d" % (SP(quelle), erste), zuordnung,
                     numerisch=numerisch)


# ===========================================================================
def baue_haltungen(pfad):
    W.setze_geometrie(14)
    W.KANTE = True
    wb = openpyxl.Workbook()
    _erzwinge_neuberechnung(wb)
    ws = wb.active
    ws.title = "Haltungen"
    W.grundgeruest(ws, BREITE_HALTUNGEN)
    erste = W.Z_DATEN

    ZK, PRUEF, SAN, KOSTEN, EIG, AD, STATUS = 10, 11, 12, 14, 15, 16, 26
    b = lambda sp: _bereich(sp, erste)

    W.block(ws, 1, 3, "Zustandsklasse",
            [("%s  %s" % (k, BEDEUTUNG[k]), S.ZUSTANDSKLASSE[k][0], S.ZUSTANDSKLASSE[k][1],
              _zaehlt(b(ZK), k, numerisch=True)) for k in "01234"],
            spalte_wert=3,
            total=("Total Haltungen", "=COUNTA(%s)" % b(2), S.FORMAT_ZAHL))

    W.block(ws, 4, 5, "Eigentümer",
            [(name, S.EIGENTUEMER[name][0], S.EIGENTUEMER[name][1],
              _zaehlt_mehrere(b(EIG), schreibweisen))
             for name, schreibweisen in EIGNER],
            spalte_wert=5, total=("Total", "=COUNTA(%s)" % b(EIG), S.FORMAT_ZAHL))

    mat_zeilen, mat_farben = _verteilungszeilen(b(4), MATERIALIEN, S.MATERIAL,
                                                S.ERSATZFARBEN)
    W.block(ws, 6, 7, "Rohrmaterial", mat_zeilen, spalte_wert=7,
            total=("Total", "=COUNTA(%s)" % b(4), S.FORMAT_ZAHL))

    W.block(ws, 8, 10, "Prüfungsresultat",
            [(text, farbe, weiss, _zaehlt_mehrere(b(PRUEF), werte))
             for text, werte, farbe, weiss in PRUEFUNGSGRUPPEN],
            spalte_wert=10, total=("Total", "=COUNTA(%s)" % b(PRUEF), S.FORMAT_ZAHL))

    W.block(ws, 11, 12, "Sanieren",
            [(s, S.SANIEREN[s][0], S.SANIEREN[s][1], _zaehlt(b(SAN), s))
             for s in ["Ja", "Nein"]],
            spalte_wert=12, total=("Total", "=COUNTA(%s)" % b(SAN), S.FORMAT_ZAHL))

    W.block(ws, 13, 14, "Ausgeführt durch  (färbt Spalte NR.)",
            [(a, S.AUSFUEHRUNG[a][0], S.AUSFUEHRUNG[a][1], _zaehlt(b(AD), a))
             for a in AUSFUEHRENDE]
            + [("noch nicht zugeteilt", GRAU, False,
                "=COUNTA(%s)-COUNTA(%s)" % (b(2), b(AD)))],
            spalte_wert=14, total=("Total", "=COUNTA(%s)" % b(2), S.FORMAT_ZAHL))

    W.block(ws, 15, 18, "Kosten nach Eigentümer",
            [("Total %s" % name, None, False,
              _summiert_mehrere(b(EIG), schreibweisen, b(KOSTEN)))
             for name, schreibweisen in EIGNER],
            spalte_wert=18, zahlenformat=S.FORMAT_CHF,
            total=("Total gesamt", "=SUM(%s)" % b(KOSTEN), S.FORMAT_CHF))

    W.block(ws, 19, 25, "Ausmass Sanierungsmassnahmen (Total)",
            [(t, None, False, "=SUM(%s)" % b(19 + i)) for i, t in enumerate(
                ["Renovierung Inliner Stk.", "Renovierung Inliner m",
                 "Anschlüsse verpressen", "Reparatur Manschette",
                 "Linerendmanschette LEM", "Reparatur Kurzliner",
                 "Erneuerung Neubau m"])],
            spalte_wert=22, zahlenformat=S.FORMAT_GEMISCHT)

    W.block(ws, 26, 27, "Sanierungsstatus",
            [("offen", S.STATUS["offen"][0], True, _zaehlt(b(STATUS), "offen")),
             ("abgeschlossen", S.STATUS["abgeschlossen"][0], True,
              _zaehlt(b(STATUS), "abgeschlossen")),
             ("noch nicht erfasst", GRAU, False,
              "=COUNTA(%s)-COUNTA(%s)" % (b(2), b(STATUS)))],
            spalte_wert=27, total=("Total", "=COUNTA(%s)" % b(2), S.FORMAT_ZAHL))

    W.logo(ws, LOGO)
    W.ausfuehrungslegende(
        ws, [(a, S.AUSFUEHRUNG[a][0], S.AUSFUEHRUNG[a][1]) for a in AUSFUEHRENDE])

    balken = [("Zustandsklassen", 3, 5), ("Sanieren Ja / Nein", 12, 2),
              ("Eigentümer", 5, 5), ("Ausgeführt durch", 14, 5),
              ("Sanierungsstatus", 27, 3)]
    farben = {
        "Zustandsklassen": [S.ZUSTANDSKLASSE[k][0] for k in "01234"],
        "Sanieren Ja / Nein": [S.SANIEREN["Ja"][0], S.SANIEREN["Nein"][0]],
        "Eigentümer": [S.EIGENTUEMER[name][0] for name, _ in EIGNER],
        "Ausgeführt durch": [S.AUSFUEHRUNG[a][0] or "FFBFBFBF" for a in AUSFUEHRENDE] + [GRAU],
        "Sanierungsstatus": [S.STATUS["offen"][0], S.STATUS["abgeschlossen"][0], GRAU],
    }
    for i, (titel, spalte, anzahl) in enumerate(balken):
        W.anteilsbalken(ws, titel,
                        dict(min_col=spalte, min_row=W.Z_WERT_VON,
                             max_row=W.Z_WERT_VON + anzahl - 1),
                        farben[titel], BALKEN_X, 0.2 + i * BALKEN_H, BALKEN_B, BALKEN_H)
    W.mengenbalken(ws, "Rohrmaterial",
                   dict(min_col=6, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + len(mat_zeilen) - 1),
                   dict(min_col=7, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + len(mat_zeilen) - 1),
                   mat_farben, MENGEN_X, 0.2, MENGEN_B, MENGEN_H)
    # Dritte Kachel: Wo liegt das Geld? Speist sich aus dem Kostenblock (SUMIF),
    # Eigner-Farben wie ueberall. Nullen bleiben dank Zahlenformat stumm.
    # Kategorien aus dem Eigentuemer-Block (Spalte 4): dort stehen die nackten
    # Namen - im Kostenblock heissen die Zeilen "Total AWU" usw.
    W.mengenbalken(ws, "Kosten nach Eigentümer",
                   dict(min_col=4, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + len(EIGNER) - 1),
                   dict(min_col=18, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + len(EIGNER) - 1),
                   [S.EIGENTUEMER[name][0] for name, _ in EIGNER],
                   DRITT_X, 0.2, DRITT_B, MENGEN_H,
                   zahlenformat=FORMAT_CHF_KURZ)

    W.titelband(ws, 27, "", [(19, 25, "Ausmass"), (26, 27, "Status")])
    W.tabellenkopf(ws, KOPF_HALTUNGEN)

    _musterzeile(
        ws, 27,
        rechts={5, 7, 14, 19, 20, 21, 22, 23, 24, 25},
        mitte={1, 10, 11, 12, 15, 16, 26, 27},
        umbruch={9, 13, 17},
        formate={7: S.FORMAT_METER, 14: S.FORMAT_CHF, 20: S.FORMAT_METER,
                 25: S.FORMAT_METER, 5: S.FORMAT_ZAHL},
        link_spalte=18)

    _farbregeln(ws, [
        (1, AD, S.AUSFUEHRUNG, False),
        (AD, AD, S.AUSFUEHRUNG, False),
        (ZK, ZK, S.ZUSTANDSKLASSE_REGEL, True),
        (PRUEF, PRUEF, S.PRUEFUNG, False),
        (SAN, SAN, S.SANIEREN, False),
        (EIG, EIG, S.EIGENTUEMER, False),
        (STATUS, STATUS, S.STATUS, False),
    ], erste)
    # Zebra zuletzt registrieren: die Bedeutungsfarben (stopIfTrue) gewinnen.
    W.zebra_regel(ws, "A%d:%s%d" % (erste, SP(27), LETZTE_ZEILE))

    W.blatt_abschluss(ws, 27, LETZTE_ZEILE)
    wb.save(pfad)


# ===========================================================================
def baue_schaechte(pfad):
    W.setze_geometrie(14)
    W.KANTE = True
    wb = openpyxl.Workbook()
    _erzwinge_neuberechnung(wb)
    ws = wb.active
    ws.title = "Schaechte"
    W.grundgeruest(ws, BREITE_SCHAECHTE)
    erste = W.Z_DATEN

    ZK, SAN, KOSTEN, EIG, AD, BELK, STATUS = 6, 7, 9, 10, 11, 15, 16
    b = lambda sp: _bereich(sp, erste)

    W.block(ws, 1, 3, "Zustandsklasse",
            [("%s  %s" % (k, BEDEUTUNG[k]), S.ZUSTANDSKLASSE[k][0], S.ZUSTANDSKLASSE[k][1],
              _zaehlt(b(ZK), k, numerisch=True)) for k in "01234"],
            spalte_wert=3,
            total=("Total Schächte", "=COUNTA(%s)" % b(3), S.FORMAT_ZAHL))

    fkt_zeilen, fkt_farben = _verteilungszeilen(
        b(2), FUNKTIONEN, {}, ["FF4472C4", "FF8497B0", "FF9DC3E6"])
    W.block(ws, 4, 6, "Funktion", fkt_zeilen, spalte_wert=6,
            total=("Total", "=COUNTA(%s)" % b(2), S.FORMAT_ZAHL))

    W.block(ws, 7, 9, "Sanieren",
            [("Ja", S.SANIEREN["Ja"][0], False, _zaehlt(b(SAN), "Ja")),
             ("Nein", S.SANIEREN["Nein"][0], False, _zaehlt(b(SAN), "Nein")),
             ("unklar (U)", "FFFFFF00", False, _zaehlt(b(SAN), "U"))],
            spalte_wert=9, total=("Kosten total", "=SUM(%s)" % b(KOSTEN), S.FORMAT_CHF))

    W.block(ws, 10, 11, "Eigentümer",
            [(name, S.EIGENTUEMER[name][0], S.EIGENTUEMER[name][1],
              _zaehlt_mehrere(b(EIG), schreibweisen))
             for name, schreibweisen in EIGNER],
            spalte_wert=11, total=("Total", "=COUNTA(%s)" % b(EIG), S.FORMAT_ZAHL))

    W.block(ws, 12, 13, "Ausgeführt durch  (färbt Spalte NR.)",
            [(a, S.AUSFUEHRUNG[a][0], S.AUSFUEHRUNG[a][1], _zaehlt(b(AD), a))
             for a in AUSFUEHRENDE]
            + [("noch nicht zugeteilt", GRAU, False,
                "=COUNTA(%s)-COUNTA(%s)" % (b(3), b(AD)))],
            spalte_wert=13, total=("Total", "=COUNTA(%s)" % b(3), S.FORMAT_ZAHL))

    W.block(ws, 14, 15, "Abdeckung / Belastungsklasse",
            [(k, None, False, _zaehlt(b(BELK), k))
             for k in ["A15", "B125", "C250", "D400", "E600", "F900"]],
            spalte_wert=15,
            total=("Abdeckungen Stk.", "=SUM(%s)" % b(14), S.FORMAT_ZAHL))

    W.block(ws, 16, 17, "Sanierungsstatus",
            [("offen", S.STATUS["offen"][0], True, _zaehlt(b(STATUS), "offen")),
             ("abgeschlossen", S.STATUS["abgeschlossen"][0], True,
              _zaehlt(b(STATUS), "abgeschlossen")),
             ("noch nicht erfasst", GRAU, False,
              "=COUNTA(%s)-COUNTA(%s)" % (b(3), b(STATUS)))],
            spalte_wert=17, total=("Total", "=COUNTA(%s)" % b(3), S.FORMAT_ZAHL))

    W.logo(ws, LOGO)
    W.ausfuehrungslegende(
        ws, [(a, S.AUSFUEHRUNG[a][0], S.AUSFUEHRUNG[a][1]) for a in AUSFUEHRENDE])

    balken = [
        ("Zustandsklassen", 3, 5, [S.ZUSTANDSKLASSE[k][0] for k in "01234"]),
        ("Sanieren Ja / Nein", 9, 3,
         [S.SANIEREN["Ja"][0], S.SANIEREN["Nein"][0], "FFFFFF00"]),
        ("Eigentümer", 11, 5, [S.EIGENTUEMER[name][0] for name, _ in EIGNER]),
        ("Ausgeführt durch", 13, 5,
         [S.AUSFUEHRUNG[a][0] or "FFBFBFBF" for a in AUSFUEHRENDE] + [GRAU]),
        ("Sanierungsstatus", 17, 3,
         [S.STATUS["offen"][0], S.STATUS["abgeschlossen"][0], GRAU]),
    ]
    for i, (titel, spalte, anzahl, farbe) in enumerate(balken):
        W.anteilsbalken(ws, titel,
                        dict(min_col=spalte, min_row=W.Z_WERT_VON,
                             max_row=W.Z_WERT_VON + anzahl - 1),
                        farbe, BALKEN_X, 0.2 + i * BALKEN_H, BALKEN_B, BALKEN_H)
    W.mengenbalken(ws, "Funktion",
                   dict(min_col=4, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + len(fkt_zeilen) - 1),
                   dict(min_col=6, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + len(fkt_zeilen) - 1),
                   fkt_farben, MENGEN_X, 0.2, MENGEN_B, MENGEN_H)
    W.mengenbalken(ws, "Abdeckungen nach Belastungsklasse",
                   dict(min_col=14, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + 5),
                   dict(min_col=15, min_row=W.Z_WERT_VON,
                        max_row=W.Z_WERT_VON + 5),
                   ["FF9DC3E6", "FF76A5AF", "FF4472C4", "FF2F5597",
                    "FF8497B0", "FF44546A"],
                   DRITT_X, 0.2, DRITT_B, MENGEN_H)

    W.titelband(ws, 17, "", [(14, 15, "Ausmass"), (16, 17, "Status")])
    W.tabellenkopf(ws, KOPF_SCHAECHTE)

    _musterzeile(
        ws, 17,
        rechts={9, 14},
        mitte={1, 3, 6, 7, 10, 11, 15, 16, 17},
        umbruch={5, 8, 12},
        formate={9: S.FORMAT_CHF, 14: S.FORMAT_ZAHL},
        link_spalte=13)

    _farbregeln(ws, [
        (1, AD, S.AUSFUEHRUNG, False),
        (AD, AD, S.AUSFUEHRUNG, False),
        (ZK, ZK, S.ZUSTANDSKLASSE_REGEL, True),
        (SAN, SAN, S.SANIEREN, False),
        (EIG, EIG, S.EIGENTUEMER, False),
        (STATUS, STATUS, S.STATUS, False),
    ], erste)
    W.zebra_regel(ws, "A%d:%s%d" % (erste, SP(17), LETZTE_ZEILE))

    W.blatt_abschluss(ws, 17, LETZTE_ZEILE)
    wb.save(pfad)


if __name__ == "__main__":
    os.makedirs(AUSGABE, exist_ok=True)
    h = os.path.join(AUSGABE, "Haltungen.xlsx")
    s = os.path.join(AUSGABE, "Schächte.xlsx")
    baue_haltungen(h)
    baue_schaechte(s)
    print("Vorlagen gebaut:")
    print("  ", h)
    print("  ", s)
    print("   Kopfzeile Zeile %d, erste Datenzeile Zeile %d, Formeln bis %d"
          % (W.Z_TABKOPF, W.Z_DATEN, LETZTE_ZEILE))
    if "--uebernehmen" in sys.argv:
        shutil.copy(h, os.path.join(ZIEL, "Haltungen.xlsx"))
        shutil.copy(s, os.path.join(ZIEL, "Schächte.xlsx"))
        print("Nach", ZIEL, "uebernommen.")
    else:
        print("Nur gebaut. Mit --uebernehmen werden die Vorlagen ersetzt.")
