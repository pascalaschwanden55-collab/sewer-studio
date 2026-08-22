# -*- coding: utf-8 -*-
"""Gemeinsames Aussehen der Excel-Berichte - genau die Farben aus Zone.1_15."""
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- Ampel- und Bedeutungsfarben (1:1 aus der Kundendatei ausgelesen) --------
ZUSTANDSKLASSE = {  # 0 = schlechteste, 4 = beste
    "0": ("FFFF0000", True,  "sofortiger Handlungsbedarf"),
    "1": ("FFFF6600", True,  "kurzfristig"),
    "2": ("FFFFFF00", False, "mittelfristig"),
    "3": ("FFAEB135", False, "langfristig"),
    "4": ("FF92D050", False, "kein Handlungsbedarf"),
}
# Alle folgenden Farben stammen aus ZustandsklasseCellStyleFactory im Programm,
# damit Tabelle in der App und Excel-Export exakt gleich aussehen.
# Gleiche Farben, aber ohne den Bedeutungstext - fuer die bedingte Formatierung,
# die nur den nackten Zellwert 0..4 vergleicht.
ZUSTANDSKLASSE_REGEL = {k: (v[0], v[1]) for k, v in ZUSTANDSKLASSE.items()}

EIGENTUEMER = {
    "AWU":      ("FF548235", True),
    "Kanton":   ("FFFFFF00", False),
    "Bund":     ("FFFF8000", True),
    "Gemeinde": ("FF00B0F0", False),
    "Privat":   ("FFFF0000", True),
}
# Feld "Ausgefuehrt_durch": faerbt im Programm die Zelle, hier zusaetzlich die
# Spalte NR. - genau wie in der Legende oben links der Datei 1.15.
# "Abwasser Uri" steht in der Legende der Datei 1.15, das Auswahlfeld im Programm
# kennt diesen Wert aber nicht - siehe Hinweis im Bericht.
AUSFUEHRUNG = {
    "Abwasser Uri":  ("FF548235", True),
    "Kanalsanierer": ("FFBF8F00", True),
    "Baumeister":    ("FF00B0F0", False),
    # Vorschlag: das Programm gibt dem Gartenbauer bisher KEINE Farbe. In einer
    # Legende waere ein leeres Feld aber ein Fehler - deshalb hier violett.
    "Gartenbauer":   ("FF7030A0", True),
}
PRUEFUNG = {
    "i.O.":                                      ("FF92D050", False),
    "beobachten":                                ("FFFFFF00", False),
    "Sanierungsbedarf":                          ("FFFF0000", True),
    "Prüfung bestanden":                      ("FF92D050", False),
    "Prüfung knapp nicht bestanden":          ("FFFFFF00", False),
    "Prüfung nicht bestanden (grob undicht)": ("FFFF0000", True),
    # Die Auswahllisten bestehender Projekte verwenden auch die ASCII-Schreibweise.
    # Das ist derselbe Textinhalt, aber kein Grund, gespeicherte Daten umzuschreiben.
    "Pruefung bestanden":                        ("FF92D050", False),
    "Pruefung knapp nicht bestanden":            ("FFFFFF00", False),
    "Pruefung nicht bestanden (grob undicht)":   ("FFFF0000", True),
    "Keine":                                  ("FFE7E6E6", False),
}
SANIEREN = {
    "Ja":   ("FF92D050", False),
    "Nein": ("FFE7E6E6", False),
}
STATUS = {
    "offen":         ("FFFF0000", True),
    "abgeschlossen": ("FF00B050", True),
}
MATERIAL = {
    "Polypropylen": ("FF70AD47", True),  "STZ": ("FFF4B183", False),
    "SBR": ("FF8497B0", False),          "Polyvinylchlorid": ("FFFF6600", True),
    # Polyethylen war reines Schwarz (aus der Kundendatei) - als Riesenbalken im
    # Diagramm erschlug das alles. Anthrazit bleibt "das dunkle Material".
    "HDPE": ("FFD9D9D9", False),         "Polyethylen": ("FF3F3F46", True),
    "Beton": ("FF808080", True),         "Zement": ("FFA6A6A6", False),
    "Steinzeug": ("FFF4B183", False),
}

# --- Grundgeruest -----------------------------------------------------------
# Moderner, ruhiger Grundton: dunkles Schiefergrau traegt Kopfzeile und
# Blocktitel, das AWU-Gruen bleibt als Akzentlinie unter dem Titel. Die alten
# Vollfarb-Baender (hellgruen/hellblau) stammten aus der 1.15 und sind bewusst
# ersetzt - die BEDEUTUNGSfarben (Ampel, Eigentuemer, Ausfuehrung) bleiben exakt.
TITEL_BAND   = "FFE9F2E3"   # zartes Gruen nur noch fuer die Ausmass-/Status-Pillen
KOPF_ZEILE   = "FF44546A"   # Kopfzeile dunkel, Schrift weiss
BLOCK_KOPF   = "FF44546A"   # Blockueberschriften im selben Ton
BLOCK_FELD   = "FFF5F7FA"
ZEBRA        = "FFF7F9FC"
RAHMEN_FEIN  = "FFD6DCE4"
RAHMEN_STARK = "FF8497B0"
SUMMEN_BAND  = "FFE2EFDA"
AKZENT       = "FF70AD47"   # AWU-Gruen als Akzent
TINTE        = "FF1F2933"
SCHIEFER_TIEF= "FF2F3B4C"
HAARLINIE    = "FFE2E7EE"

SCHRIFT = "Calibri"

def fuellung(argb=None):
    return PatternFill("solid", fgColor=argb) if argb else PatternFill()

def schrift(groesse=9, fett=False, weiss=False, farbe=None, kursiv=False, unterstrichen=None):
    return Font(name=SCHRIFT, size=groesse, bold=fett, italic=kursiv,
                color=("FFFFFFFF" if weiss else (farbe or "FF1F2933")),
                underline=unterstrichen)

def ausrichtung(h="left", v="center", umbruch=False, einzug=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=umbruch, indent=einzug)

def _kante(farbe, stil="thin"):
    return Side(style=stil, color=farbe)

GITTER      = Border(left=_kante(RAHMEN_FEIN), right=_kante(RAHMEN_FEIN),
                     top=_kante(RAHMEN_FEIN), bottom=_kante(RAHMEN_FEIN))
GITTER_KOPF = Border(left=_kante(RAHMEN_STARK), right=_kante(RAHMEN_STARK),
                     top=_kante(RAHMEN_STARK), bottom=_kante(RAHMEN_STARK, "medium"))
BLOCK_RAHMEN= Border(left=_kante(RAHMEN_STARK), right=_kante(RAHMEN_STARK),
                     top=_kante(RAHMEN_STARK), bottom=_kante(RAHMEN_STARK))

# Datenzeilen: feine Linien in beide Richtungen. Ganz ohne senkrechte Striche
# verliert das Auge bei 27 Spalten die Spur - die Zeile darunter waagrecht zu
# halten reicht nicht. Die Spaltentrenner sind heller als die Zeilentrenner,
# damit die Tabelle trotzdem ruhig bleibt und nicht wie ein Karogitter wirkt.
# Ein durchgehendes, klar sichtbares Gitter. Zeilen und Spalten tragen dieselbe
# Staerke - bei 27 Spalten und unterschiedlich hohen Zeilen (die Massnahmen
# bestimmen die Hoehe) braucht das Auge beide Richtungen als Fuehrung. Der Ton
# ist ein ruhiges Blaugrau, kein Schwarz: sichtbar, aber nicht laut.
GITTERLINIE = "FFB4C0CE"
SPALTENLINIE = GITTERLINIE
ZEILENLINIE = Border(bottom=_kante(GITTERLINIE), top=_kante(GITTERLINIE),
                     left=_kante(GITTERLINIE), right=_kante(GITTERLINIE))

# Kopfzeile: dunkle Flaeche, Spalten durch einen helleren Ton getrennt - so
# sieht man auch im Kopf, wo eine Spalte endet. Unten eine kraeftige Kante.
KOPF_TRENNER = "FF5D6E85"
KOPF_RAHMEN = Border(left=_kante(KOPF_TRENNER), right=_kante(KOPF_TRENNER),
                     top=_kante(KOPF_ZEILE), bottom=_kante(SCHIEFER_TIEF, "medium"))

# Titelzeile: weiss, darunter die gruene Akzentlinie.
TITEL_KANTE = Border(bottom=_kante(AKZENT, "thick"))

FORMAT_CHF   = '"CHF" #,##0.00'
FORMAT_METER = '#,##0.00'
FORMAT_ZAHL  = '#,##0'
FORMAT_GEMISCHT = "#,##0.##"


# Reservefarben fuer Werte ohne benannte Farbe (das Programm kennt 17
# Rohrmaterialien, benannte Farben gibt es nur fuer die haeufigsten).
ERSATZFARBEN = ["FF9DC3E6", "FFFFD966", "FFB4A7D6", "FF76A5AF", "FFE06666", "FFC9C9C9"]


def weisse_schrift(argb):
    """Dunkler Hintergrund bekommt weisse Schrift, damit alles lesbar bleibt."""
    if not argb:
        return False
    r, g, b = int(argb[-6:-4], 16), int(argb[-4:-2], 16), int(argb[-2:], 16)
    return (0.299 * r + 0.587 * g + 0.114 * b) < 150
