# -*- coding: utf-8 -*-
"""Bausteine fuer den Excel-Bericht: Kopfzone, Kennzahlenbloecke, Tabelle."""
from openpyxl.utils import get_column_letter as SP
from openpyxl.utils.units import cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList, DataLabel
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (Paragraph, ParagraphProperties,
                                   CharacterProperties, RichTextProperties)
from openpyxl.drawing.line import LineProperties
from openpyxl.worksheet.properties import PageSetupProperties
import stil as S


def _kleine_schrift(punkt=7, fett=False):
    """Kleine Schrift fuer Diagrammlegende, Titel und Beschriftungen."""
    zeichen = CharacterProperties(sz=punkt * 100, b=fett)
    return RichText(bodyPr=RichTextProperties(),
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=zeichen),
                                 endParaRPr=zeichen)])

# Geometrie: gilt fuer beide Blaetter, damit sie gleich aussehen
# Kopfzone: Zeilen 1-4 Logo, 5-10 Legende "Ausgefuehrt durch" (oben links wie in
# der Datei 1.15), rechts daneben die Diagrammkacheln.
Z_KOPFZONE = 11
Z_LEGENDE_KOPF = 5
Z_LEGENDE_VON = 6
Z_TRENNER_1 = 12
Z_BLOCKKOPF = 13
Z_WERT_VON = 14
Z_WERT_BIS = 19
Z_TOTAL = 20
Z_TRENNER_2 = 21
Z_TITEL = 22
Z_TABKOPF = 23
Z_DATEN = 24

# Ruhiger Stil: Farbe als linke Kante statt als ganze Fuellung.
KANTE = False


def setze_geometrie(kopfzone):
    """Verschiebt alle Zeilenmarken, wenn die Kopfzone hoeher wird."""
    global Z_KOPFZONE, Z_TRENNER_1, Z_BLOCKKOPF, Z_WERT_VON, Z_WERT_BIS
    global Z_TOTAL, Z_TRENNER_2, Z_TITEL, Z_TABKOPF, Z_DATEN
    Z_KOPFZONE = kopfzone
    Z_TRENNER_1 = kopfzone + 1
    Z_BLOCKKOPF = kopfzone + 2
    Z_WERT_VON = kopfzone + 3
    Z_WERT_BIS = kopfzone + 8
    Z_TOTAL = kopfzone + 9
    Z_TRENNER_2 = kopfzone + 10
    Z_TITEL = kopfzone + 11
    Z_TABKOPF = kopfzone + 12
    Z_DATEN = kopfzone + 13


def grundgeruest(ws, breiten):
    """Spaltenbreiten, Zeilenhoehen und Blatteinstellungen."""
    for i, b in enumerate(breiten, start=1):
        ws.column_dimensions[SP(i)].width = b
    for r in range(1, Z_KOPFZONE + 1):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[Z_TRENNER_1].height = 5
    ws.row_dimensions[Z_BLOCKKOPF].height = 20
    for r in range(Z_WERT_VON, Z_WERT_BIS + 1):
        ws.row_dimensions[r].height = 16
    ws.row_dimensions[Z_TOTAL].height = 17
    ws.row_dimensions[Z_TRENNER_2].height = 5
    ws.row_dimensions[Z_TITEL].height = 30
    ws.row_dimensions[Z_TABKOPF].height = 34
    ws.sheet_view.showGridLines = False


def logo(ws, pfad, x_cm=0.6, y_cm=0.15, breite_cm=5.6):
    bild = Image(pfad)
    hoehe_cm = breite_cm * bild.height / bild.width
    bild.anchor = AbsoluteAnchor(
        pos=XDRPoint2D(cm_to_EMU(x_cm), cm_to_EMU(y_cm)),
        ext=XDRPositiveSize2D(cm_to_EMU(breite_cm), cm_to_EMU(hoehe_cm)))
    ws.add_image(bild)


def kreisdiagramm(ws, titel, bereich_beschriftung, bereich_werte, farben,
                  x_cm, y_cm, breite_cm=6.4, hoehe_cm=5.0, werte=None,
                  mindestanteil=0.05):
    """Ein Kreisdiagramm mit fest zugeordneten Ampelfarben je Segment."""
    d = PieChart()
    d.title = titel
    d.height = hoehe_cm
    d.width = breite_cm
    d.add_data(Reference(ws, **bereich_werte), titles_from_data=False)
    d.set_categories(Reference(ws, **bereich_beschriftung))
    reihe = d.series[0]
    for i, f in enumerate(farben):
        punkt = DataPoint(idx=i)
        punkt.graphicalProperties = GraphicalProperties(solidFill=f[-6:])
        punkt.graphicalProperties.line = LineProperties(solidFill="FFFFFF", w=12700)
        reihe.data_points.append(punkt)
    d.dataLabels = DataLabelList()
    d.dataLabels.showPercent = True
    d.dataLabels.showVal = False
    d.dataLabels.showCatName = False
    d.dataLabels.showSerName = False
    d.dataLabels.showLegendKey = False
    d.dataLabels.dLblPos = "outEnd"
    d.dataLabels.txPr = _kleine_schrift(7)

    # Winzige Segmente bekommen keine Prozentzahl: zwei Beschriftungen mit 2 %
    # liegen sonst uebereinander und beide sind unlesbar. Die genaue Zahl steht
    # ohnehin im Kennzahlenblock darunter.
    if werte:
        gesamt = sum(werte) or 1
        stumm = []
        for i, v in enumerate(werte):
            if v / gesamt < mindestanteil:
                einzeln = DataLabel(idx=i)
                einzeln.showPercent = False
                einzeln.showVal = False
                einzeln.showCatName = False
                einzeln.showSerName = False
                einzeln.showLegendKey = False
                einzeln.showBubbleSize = False
                stumm.append(einzeln)
        if stumm:
            d.dataLabels.dLbl = stumm

    d.legend.position = "b"
    d.legend.overlay = False
    d.legend.txPr = _kleine_schrift(7)
    if d.title is not None:
        d.title.tx.rich.p[0].pPr = ParagraphProperties(
            defRPr=CharacterProperties(sz=1000, b=True))
        for lauf in d.title.tx.rich.p[0].r or []:
            lauf.rPr = CharacterProperties(sz=1000, b=True)
    try:
        rahmen = GraphicalProperties(solidFill="FFFFFF")
        rahmen.line = LineProperties(solidFill=S.RAHMEN_FEIN[-6:], w=9525)
        d.graphical_properties = rahmen
    except Exception:
        pass
    d.anchor = AbsoluteAnchor(
        pos=XDRPoint2D(cm_to_EMU(x_cm), cm_to_EMU(y_cm)),
        ext=XDRPositiveSize2D(cm_to_EMU(breite_cm), cm_to_EMU(hoehe_cm)))
    ws.add_chart(d)


def ausfuehrungslegende(ws, eintraege, spalte_chip=1, spalte_name=2, spalte_name_bis=3):
    """Legende oben links - sie erklaert nur die Farben, sie zaehlt nichts.

    Aufbau wie in der Datei 1.15: ein farbiges "Nr." (so wie die Zelle in der
    Spalte NR. aussieht) und daneben, wer ausfuehrt. Die Zahlen dazu stehen im
    Kennzahlenband unten.

    eintraege: Liste aus (Name, Fuellfarbe oder None, weisse_Schrift)
    """
    ws.merge_cells(start_row=Z_LEGENDE_KOPF, start_column=spalte_chip,
                   end_row=Z_LEGENDE_KOPF, end_column=spalte_name_bis)
    k = ws.cell(Z_LEGENDE_KOPF, spalte_chip)
    k.value = "Ausgeführt durch  –  Farbe der Spalte NR."
    k.fill = S.fuellung(S.BLOCK_KOPF)
    k.font = S.schrift(9, fett=True, weiss=True)
    k.alignment = S.ausrichtung("center")

    for i, (name, farbe, weiss) in enumerate(eintraege):
        r = Z_LEGENDE_VON + i
        chip = ws.cell(r, spalte_chip)
        chip.value = "Nr."
        chip.fill = S.fuellung(farbe) if farbe else S.fuellung()
        chip.font = S.schrift(9, fett=True, weiss=weiss)
        chip.alignment = S.ausrichtung("center")

        if spalte_name_bis > spalte_name:
            ws.merge_cells(start_row=r, start_column=spalte_name,
                           end_row=r, end_column=spalte_name_bis)
        bez = ws.cell(r, spalte_name)
        bez.value = name
        for c in range(spalte_name, spalte_name_bis + 1):
            ws.cell(r, c).fill = S.fuellung(S.BLOCK_FELD)
        bez.font = S.schrift(9, fett=True)
        bez.alignment = S.ausrichtung("left")

    letzte = Z_LEGENDE_VON + len(eintraege) - 1
    for r in range(Z_LEGENDE_KOPF, letzte + 1):
        for c in range(spalte_chip, spalte_name_bis + 1):
            ws.cell(r, c).border = S.GITTER
    return letzte


def farbregeln(ws, ziel_bereich, quelle_zelle, zuordnung, numerisch=False):
    """Bedingte Formatierung statt fester Farbe.

    Damit folgt die Farbe dem Wert auch dann, wenn der Wert spaeter in Excel von
    Hand geaendert wird. Eine feste Fuellung wuerde stehenbleiben und luegen.

    zuordnung: {Zellwert: (ARGB oder None, weisse_Schrift)}
    """
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill, Font

    for wert, (argb, weiss) in zuordnung.items():
        if not argb:
            continue
        if numerisch:
            # Zwei Fallen auf einmal:
            # 1. Ohne die Leerpruefung faerbt "= 0" jede LEERE Zelle - Excel
            #    liest eine leere Zelle im Zahlenvergleich als 0.
            # 2. Der Export schreibt die Zustandsklasse als TEXT, von Hand
            #    getippt waere sie eine Zahl. Excel wandelt beim Vergleich
            #    nicht um, deshalb muessen beide Formen geprueft werden.
            formel = 'AND(%s<>"",OR(%s=%s,%s="%s"))' % (
                quelle_zelle, quelle_zelle, wert, quelle_zelle, wert)
        else:
            formel = '%s="%s"' % (quelle_zelle, wert)
        ws.conditional_formatting.add(ziel_bereich, FormulaRule(
            formula=[formel],
            fill=PatternFill(start_color=argb, end_color=argb, fill_type="solid"),
            font=Font(name=S.SCHRIFT, size=9, bold=True,
                      color="FFFFFFFF" if weiss else "FF1F2933"),
            stopIfTrue=True))


def _anker(x_cm, y_cm, breite_cm, hoehe_cm):
    return AbsoluteAnchor(
        pos=XDRPoint2D(cm_to_EMU(x_cm), cm_to_EMU(y_cm)),
        ext=XDRPositiveSize2D(cm_to_EMU(breite_cm), cm_to_EMU(hoehe_cm)))


def _schmucklos(diagramm):
    """Kein Rahmen, keine Gitterlinien - moderne, ruhige Darstellung."""
    try:
        flaeche = GraphicalProperties(solidFill="FFFFFF")
        flaeche.line = LineProperties(noFill=True)
        diagramm.graphical_properties = flaeche
    except Exception:
        pass


def _titel_klein(diagramm, punkt=10):
    if diagramm.title is not None:
        diagramm.title.tx.rich.p[0].pPr = ParagraphProperties(
            defRPr=CharacterProperties(sz=punkt * 100, b=True))
        for lauf in diagramm.title.tx.rich.p[0].r or []:
            lauf.rPr = CharacterProperties(sz=punkt * 100, b=True)


def anteilsbalken(ws, titel, bereich_werte, farben, x_cm, y_cm,
                  breite_cm=12.0, hoehe_cm=1.9):
    """Ein einzelner gestapelter 100-Prozent-Balken, waagrecht.

    Fuer geordnete Groessen wie die Zustandsklasse ist das die richtige Form:
    der Balken behaelt die Reihenfolge 0..4 bei, ein Kuchen zerstoert sie.
    Jede Zeile des Kennzahlenblocks wird eine eigene Reihe mit einem Wert.
    """
    d = BarChart()
    d.type = "bar"
    d.grouping = "percentStacked"
    d.overlap = 100
    d.title = titel
    d.width = breite_cm
    d.height = hoehe_cm
    d.add_data(Reference(ws, **bereich_werte), from_rows=True, titles_from_data=False)

    for reihe, farbe in zip(d.series, farben):
        eigenschaften = GraphicalProperties(solidFill=farbe[-6:])
        eigenschaften.line = LineProperties(solidFill="FFFFFF", w=9525)
        reihe.graphicalProperties = eigenschaften

    d.dataLabels = DataLabelList()
    d.dataLabels.showVal = True
    d.dataLabels.showSerName = False
    d.dataLabels.showCatName = False
    d.dataLabels.showLegendKey = False
    # Nullen bleiben stumm: ein leeres Segment braucht keine "0" im Bild.
    d.dataLabels.numFmt = '#,##0;-#,##0;""'
    d.dataLabels.txPr = _kleine_schrift(7, fett=True)

    d.legend = None
    d.y_axis.delete = True
    d.x_axis.delete = True
    d.gapWidth = 20
    _schmucklos(d)
    _titel_klein(d, 9)
    d.anchor = _anker(x_cm, y_cm, breite_cm, hoehe_cm)
    ws.add_chart(d)


def mengenbalken(ws, titel, bereich_beschriftung, bereich_werte, farben,
                 x_cm, y_cm, breite_cm=8.0, hoehe_cm=6.0,
                 zahlenformat='#,##0;-#,##0;""'):
    """Liegendes Balkendiagramm - vertraegt beliebig viele Auspraegungen.

    Fuer das Rohrmaterial (bis 17 Arten) ist ein Kuchen unlesbar; nebeneinander
    liegende Balken lassen sich dagegen direkt vergleichen.
    """
    d = BarChart()
    d.type = "bar"
    d.grouping = "clustered"
    d.title = titel
    d.width = breite_cm
    d.height = hoehe_cm
    d.add_data(Reference(ws, **bereich_werte), titles_from_data=False)
    d.set_categories(Reference(ws, **bereich_beschriftung))

    reihe = d.series[0]
    # Die Beschriftungen sind Text, nicht Zahlen. Ohne strRef liest der Leser sie
    # als Zahlenreihe und das Diagramm bricht (real gesehen: nur ein Balken).
    reihe.cat = AxDataSource(
        strRef=StrRef(f=str(Reference(ws, **bereich_beschriftung))))

    for i, farbe in enumerate(farben):
        punkt = DataPoint(idx=i)
        punkt.graphicalProperties = GraphicalProperties(solidFill=farbe[-6:])
        reihe.data_points.append(punkt)

    d.dataLabels = DataLabelList()
    d.dataLabels.showVal = True
    d.dataLabels.showCatName = False
    d.dataLabels.showSerName = False
    d.dataLabels.showLegendKey = False
    d.dataLabels.showPercent = False
    d.dataLabels.showBubbleSize = False
    # Nullen bleiben stumm - sonst steht an jedem leeren Balken eine "0".
    d.dataLabels.numFmt = zahlenformat
    d.dataLabels.txPr = _kleine_schrift(7)
    d.legend = None
    d.gapWidth = 40
    # Achtung: bei openpyxl ist x_axis IMMER die Kategorienachse, auch beim
    # liegenden Balken. Die Namen muessen bleiben, die Werteachse kann weg -
    # die Zahlen stehen als Beschriftung am Balken.
    d.x_axis.delete = False
    d.x_axis.majorGridlines = None
    d.x_axis.txPr = _kleine_schrift(7)
    d.y_axis.delete = True
    d.y_axis.majorGridlines = None
    _schmucklos(d)
    _titel_klein(d, 9)
    d.anchor = _anker(x_cm, y_cm, breite_cm, hoehe_cm)
    ws.add_chart(d)


def _rahmen_setzen(ws, z1, s1, z2, s2, rahmen):
    for r in range(z1, z2 + 1):
        for c in range(s1, s2 + 1):
            ws.cell(r, c).border = rahmen


def _kantenrahmen(argb):
    """Farbe als kraeftige linke Kante statt als ganze Fuellung - ruhiger Look."""
    from openpyxl.styles import Border, Side
    return Border(left=Side(style="thick", color=argb),
                  right=Side(style="thin", color=S.RAHMEN_FEIN),
                  top=Side(style="thin", color=S.RAHMEN_FEIN),
                  bottom=Side(style="thin", color=S.RAHMEN_FEIN))


def block(ws, spalte_von, spalte_bis, kopf, zeilen, spalte_wert=None,
          zahlenformat=S.FORMAT_ZAHL, total=None, kante=None):
    """Ein Kennzahlenblock: Ueberschrift, farbige Zeilen, Totalzeile.

    zeilen: Liste aus (Beschriftung, Fuellfarbe oder None, weisse_Schrift, Wert oder None)
    Ohne Totalzeile darf der Block eine Zeile mehr belegen.
    """
    if kante is None:
        kante = KANTE
    if spalte_wert is None:
        spalte_wert = spalte_bis
    spalte_text_bis = spalte_wert - 1 if spalte_wert > spalte_von else spalte_von
    letzte_wertzeile = Z_WERT_BIS if total is not None else Z_TOTAL

    ws.merge_cells(start_row=Z_BLOCKKOPF, start_column=spalte_von,
                   end_row=Z_BLOCKKOPF, end_column=spalte_bis)
    k = ws.cell(Z_BLOCKKOPF, spalte_von)
    k.value = kopf
    k.fill = S.fuellung(S.BLOCK_KOPF)
    k.font = S.schrift(9, fett=True, weiss=True)
    k.alignment = S.ausrichtung("center")

    for i, (text, farbe, weiss, wert) in enumerate(zeilen):
        r = Z_WERT_VON + i
        if r > letzte_wertzeile:
            break
        if spalte_text_bis > spalte_von:
            ws.merge_cells(start_row=r, start_column=spalte_von,
                           end_row=r, end_column=spalte_text_bis)
        t = ws.cell(r, spalte_von)
        t.value = text
        if kante:
            t.fill = S.fuellung(S.BLOCK_FELD)
            t.font = S.schrift(9)
            t.alignment = S.ausrichtung("left", einzug=1)
        else:
            t.fill = S.fuellung(farbe) if farbe else S.fuellung(S.BLOCK_FELD)
            t.font = S.schrift(9, weiss=weiss, fett=bool(farbe))
            t.alignment = S.ausrichtung("center" if spalte_text_bis == spalte_von else "left")
        if wert is not None:
            if spalte_bis > spalte_wert:
                ws.merge_cells(start_row=r, start_column=spalte_wert,
                               end_row=r, end_column=spalte_bis)
            w = ws.cell(r, spalte_wert)
            w.value = wert
            w.fill = S.fuellung(S.BLOCK_FELD)
            w.font = S.schrift(9, fett=True)
            w.alignment = S.ausrichtung("center")
            w.number_format = zahlenformat

    if total is not None:
        text, wert, fmt = total
        if spalte_text_bis > spalte_von:
            ws.merge_cells(start_row=Z_TOTAL, start_column=spalte_von,
                           end_row=Z_TOTAL, end_column=spalte_text_bis)
        t = ws.cell(Z_TOTAL, spalte_von)
        t.value = text
        t.fill = S.fuellung(S.SUMMEN_BAND)
        t.font = S.schrift(9, fett=True)
        t.alignment = S.ausrichtung("left" if spalte_text_bis > spalte_von else "center")
        if spalte_bis > spalte_wert:
            ws.merge_cells(start_row=Z_TOTAL, start_column=spalte_wert,
                           end_row=Z_TOTAL, end_column=spalte_bis)
        w = ws.cell(Z_TOTAL, spalte_wert)
        w.value = wert
        w.fill = S.fuellung(S.SUMMEN_BAND)
        w.font = S.schrift(9, fett=True)
        w.alignment = S.ausrichtung("center")
        w.number_format = fmt or zahlenformat

    # Nicht belegte Zeilen bleiben nicht weiss stehen, sonst wirkt der Block zerissen
    for r in range(Z_WERT_VON + len(zeilen), letzte_wertzeile + 1):
        for c in range(spalte_von, spalte_bis + 1):
            ws.cell(r, c).fill = S.fuellung(S.BLOCK_FELD)

    _rahmen_setzen(ws, Z_BLOCKKOPF, spalte_von, Z_TOTAL, spalte_bis, S.GITTER)
    for c in range(spalte_von, spalte_bis + 1):
        ws.cell(Z_BLOCKKOPF, c).border = S.BLOCK_RAHMEN

    # Im ruhigen Stil traegt eine kraeftige linke Kante die Farbe.
    if kante:
        for i, (_, farbe, _, _) in enumerate(zeilen):
            r = Z_WERT_VON + i
            if r <= letzte_wertzeile and farbe:
                ws.cell(r, spalte_von).border = _kantenrahmen(farbe)


def titelband(ws, spalten, titel, bereiche):
    """Titelzeile: grosser dunkler Titel auf Weiss, darunter die gruene
    Akzentlinie. Die Bereichsbaender Ausmass/Status bleiben als zarte gruene
    Pillen - das frueher vollflaechig gruene Band war der Alt-Excel-Look."""
    belegt = set()
    for von, bis, text in bereiche:
        belegt.update(range(von, bis + 1))
        ws.merge_cells(start_row=Z_TITEL, start_column=von, end_row=Z_TITEL, end_column=bis)
        z = ws.cell(Z_TITEL, von)
        z.value = text
        z.font = S.schrift(10, fett=True)
        z.alignment = S.ausrichtung("center", "bottom")
    ende = min(belegt) - 1 if belegt else spalten
    ws.merge_cells(start_row=Z_TITEL, start_column=1, end_row=Z_TITEL, end_column=ende)
    z = ws.cell(Z_TITEL, 1)
    z.value = titel
    z.font = S.schrift(14, fett=True)
    z.alignment = S.ausrichtung("left", "bottom")
    for c in range(1, spalten + 1):
        zelle = ws.cell(Z_TITEL, c)
        zelle.border = S.TITEL_KANTE
        if c in belegt:
            zelle.fill = S.fuellung(S.TITEL_BAND)


def tabellenkopf(ws, kopftexte):
    for i, t in enumerate(kopftexte, start=1):
        z = ws.cell(Z_TABKOPF, i)
        z.value = t
        z.fill = S.fuellung(S.KOPF_ZEILE)
        z.font = S.schrift(9, fett=True, weiss=True)
        z.alignment = S.ausrichtung("center", "center", umbruch=True)
        z.border = S.KOPF_RAHMEN


def zebra_regel(ws, bereich):
    """Sehr helles Zebra ueber bedingte Formatierung, NACH den Bedeutungsfarben
    registriert: deren stopIfTrue gewinnt, das Zebra fuellt nur die Luecken.
    Als Regel statt als feste Fuellung, damit der Export weiterhin keine Zelle
    selbst einfaerbt."""
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    ws.conditional_formatting.add(bereich, FormulaRule(
        formula=["MOD(ROW(),2)=0"],
        fill=PatternFill(start_color=S.ZEBRA, end_color=S.ZEBRA, fill_type="solid")))


def blatt_abschluss(ws, spalten, letzte_datenzeile):
    ws.freeze_panes = ws.cell(Z_DATEN, 1)
    ws.auto_filter.ref = "A%d:%s%d" % (Z_TABKOPF, SP(spalten), letzte_datenzeile)
    ws.print_title_rows = "%d:%d" % (Z_TITEL, Z_TABKOPF)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8  # A3
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.oddFooter.right.text = "Seite &P von &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.left.text = "&F"
    ws.oddFooter.left.size = 8
