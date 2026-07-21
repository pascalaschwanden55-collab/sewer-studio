#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Amtsblatt-Uri-Scanner  v2
-------------------------
Erfasst aus dem Volltext EINER Amtsblatt-Ausgabe:

 A) ALLE Baugesuche / Bauplanauflagen (Neubau, Anbau, Umbau, Erweiterung ...),
    weil daran Kanalisation haengen kann.
    AUSNAHME: reine Solaranlagen und reine Waermepumpen werden ausgeschlossen.
    (Ein Neubau MIT Waermepumpe bleibt drin — nur die reine Anlage fliegt raus.)

 B) Alles mit Abwasser-Bezug in der ganzen Ausgabe, egal in welcher Rubrik:
    Kanalisation, Abwasser, Entwaesserung, ARA, Klaeranlage, dezentrale
    Kleinklaeranlagen, Klaergrube usw.

Ausgabe: eine Excel-Datei pro Ausgabe.

Aufruf:
  python3 amtsblatt_scan.py INPUT.txt OUTPUT.xlsx --nr 29 --datum 17.07.2026 --url https://...
"""
import re, argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- Stichwoerter
# Abwasser-/Klaeranlagen-Bezug (Substring, case-insensitive)
ABWASSER = ["kanalisation", "abwasser", "schmutzwasser", "mischwasser",
            "meteorwasser", "regenabwasser", "entwässer", "entwaesser",
            "sammelkanal", "abwasserkanal", "kläranlage", "klaeranlage",
            "kleinkläranlage", "kleinklaeranlage", "abwasserreinigung",
            "hauskläranlage", "hausklaeranlage", "klärgrube", "klaergrube",
            "faulgrube", "pflanzenkläranlage", "klärschlamm", "klaerschlamm",
            "sickerleitung", "versickerung", "einleitbewilligung",
            "gewässerschutz", "gewaesserschutz", "dezentrale abwasser"]
ARA_RX = re.compile(r"\bARA\b")          # Ganzwort, sonst zu viele Fehltreffer

# Breitere Liste NUR zum Markieren von Baugesuchen (Priorisierung).
# Ausserhalb der Baugesuche waeren diese Begriffe zu unscharf.
BAU_FLAG = ABWASSER + [
    "sauberwasser", "einleitung", "schacht", "schächte", "schaechte",
    "leitung", "versickerung", "drainage", "werkleitung", "hausanschluss",
    "grube", "meteor", "eindolung", "rohr"]

# Ausschluss: reine Solaranlagen / Waermepumpen
SOLAR_WP = re.compile(
    r"(solaranlage|solarpanel|photovoltaik|pv-anlage|\bpv\b|\bsolar\b|"
    r"wärmepumpe|waermepumpe|erdsonde|erdwärmesonde|erdwaermesonde)", re.I)
# ... aber nur, wenn sonst kein echtes Bauobjekt genannt ist
BAUOBJEKT = re.compile(
    r"(wohnhaus|einfamilienhaus|mehrfamilienhaus|\befh\b|\bmfh\b|gebäude|"
    r"\bhaus\b|halle|stall|scheune|garage|carport|überdachung|anbau|umbau|"
    r"umnutzung|ersatzneubau|erweiterung|abbruch|terrasse|pool|mauer|zaun|"
    r"reklame|antenne|strasse|leitung|kanal|schacht|parkplatz|unterstand|"
    r"tunnel|brücke|deponie|silo|remise|werkstatt|laden|restaurant)", re.I)

MUNI = ["Altdorf", "Andermatt", "Attinghausen", "Bürglen", "Bauen", "Erstfeld",
        "Flüelen", "Göschenen", "Gurtnellen", "Hospental", "Isenthal", "Realp",
        "Schattdorf", "Seedorf", "Seelisberg", "Silenen", "Sisikon", "Spiringen",
        "Unterschächen", "Wassen"]
MUNI_SET = set(MUNI)
MUNI_RX = re.compile(r"^(?:Betroffene\s+)?Gemeinde\s+([A-Za-zÄÖÜäöüß]+)$")

SECTIONS = {
 "regierungsrat": "Regierungsrat", "direktionen": "Direktionen",
 "volkswirtschaftsdirektion": "Volkswirtschaftsdirektion", "baudirektion": "Baudirektion",
 "gesundheits-, sozial- und umweltdirektion": "Gesundheits-, Sozial- und Umweltdirektion",
 "sicherheitsdirektion": "Sicherheitsdirektion", "finanzdirektion": "Finanzdirektion",
 "bildungs- und kulturdirektion": "Bildungs- und Kulturdirektion",
 "justizdirektion": "Justizdirektion",
 "weitere behörden und einrichtungen": "Weitere Behörden und Einrichtungen",
 "eigentumsübertragungen": "Eigentumsübertragungen", "handelsregister": "Handelsregister",
 "bau- und planungsrecht": "Bau- und Planungsrecht",
 "auflage- und einspracheverfahren": "Auflage- und Einspracheverfahren",
 "bauplanauflagen": "Bauplanauflagen", "baugesuche": "Baugesuche",
 "konzession; gesuch": "Konzession; Gesuch", "konzession": "Konzession",
 "öffentliche auflage": "Öffentliche Auflage",
 "verkehrsbeschränkungen": "Verkehrsbeschränkungen", "signalisation": "Signalisation",
 "submissionen": "Submissionen", "ausschreibung": "Ausschreibungen",
 "wasserrechtsverleihungen": "Wasserrechtsverleihungen", "gerichte": "Gerichte",
 "schuldbetreibung und konkurs": "Schuldbetreibung und Konkurs",
 "rechtsauskunft": "Rechtsauskunft", "veranstaltungen": "Veranstaltungen",
}
BAU_SECTIONS = {"Bauplanauflagen", "Baugesuche"}

PAGE_RX = re.compile(r"^\s*(\d{2,4})\s*(?:Administrativer|Gerichtlicher)")
PAGE_NUM_RX = re.compile(r"^\s*(\d{2,4})\s*$")
LABELS = [("Bauherrschaft", "bauherrschaft"), ("Bauvorhaben", "bauvorhaben"),
          ("Bauplatz", "bauplatz"), ("Bemerkungen", "bemerkungen")]
# Achtung: die PDF-Aufzaehlungszeichen kommen als Buchstaben ("n n") vor dem
# Label an -> Label darf NICHT am Zeilenanfang verankert sein (sonst verschmelzen
# Eintraege und "Bauherrschaft" bleibt leer).
LABEL_RX = re.compile(r"(Bauherrschaft|Bauvorhaben|Bauplatz|Bemerkungen)\s*:\s*(.*)$")


def clean(s):
    return re.sub(r"\s+", " ", s or "").strip()


def has_abwasser(txt):
    low = (txt or "").lower()
    for kw in ABWASSER:
        if kw in low:
            return kw
    if ARA_RX.search(txt or ""):
        return "ARA"
    return None


def has_bau_flag(txt):
    """Breitere Markierung innerhalb der Baugesuche (nur zur Priorisierung)."""
    low = (txt or "").lower()
    for kw in BAU_FLAG:
        if kw in low:
            return kw
    if ARA_RX.search(txt or ""):
        return "ARA"
    return None


def is_solar_wp_only(bauvorhaben):
    """True, wenn das Bauvorhaben eine reine Solar-/Waermepumpen-Anlage ist."""
    if not bauvorhaben or not SOLAR_WP.search(bauvorhaben):
        return False
    return not BAUOBJEKT.search(bauvorhaben)


def scan(text):
    lines = text.splitlines()
    page = section = gem = ""
    bau, treffer, excluded = [], [], []
    entry = None
    field = None

    def flush():
        nonlocal entry, field
        if entry and (entry.get("bauvorhaben") or entry.get("bauherrschaft")):
            for k in ("bauherrschaft", "bauvorhaben", "bauplatz", "bemerkungen"):
                entry[k] = clean(entry.get(k, ""))
            volltext = " ".join(entry.get(k, "") for k in
                                ("bauherrschaft", "bauvorhaben", "bauplatz", "bemerkungen"))
            entry["kw"] = has_bau_flag(volltext)
            if is_solar_wp_only(entry["bauvorhaben"]):
                excluded.append(entry)
            else:
                bau.append(entry)
        entry, field = None, None

    for i, raw in enumerate(lines):
        line = raw.rstrip()
        s = line.strip()
        m = PAGE_RX.match(line) or PAGE_NUM_RX.match(line)
        if m:
            page = m.group(1)
        key = s.lower().rstrip(":")
        if key in SECTIONS:
            flush()
            section, gem = SECTIONS[key], ""
            continue
        if s in MUNI_SET:
            flush(); gem = s; continue
        mm = MUNI_RX.match(s)
        if mm and mm.group(1) in MUNI_SET:
            flush(); gem = mm.group(1); continue

        # ---- A) strukturierte Baugesuch-Eintraege
        if section in BAU_SECTIONS:
            lm = LABEL_RX.search(line)
            if lm:
                label, val = lm.group(1), lm.group(2)
                if label == "Bauherrschaft":
                    flush()
                    entry = {"gemeinde": gem, "seite": page, "rubrik": section}
                if entry is None:
                    entry = {"gemeinde": gem, "seite": page, "rubrik": section}
                field = dict(LABELS)[label]
                entry[field] = (entry.get(field, "") + " " + val).strip()
                continue
            if entry is not None and field and s:
                entry[field] = (entry.get(field, "") + " " + s).strip()
                continue

        # ---- B) Abwasser-Treffer ausserhalb der Baugesuch-Bloecke
        unit = line
        if line.endswith("-") and i + 1 < len(lines):
            unit = line[:-1] + lines[i + 1].strip()
        kw = has_abwasser(unit)
        if kw and section not in BAU_SECTIONS:
            win = clean(" ".join(lines[max(0, i - 1):i + 3]))
            treffer.append({"rubrik": section or "—", "gemeinde": gem, "seite": page,
                            "kw": kw, "auszug": win[:320]})
    flush()

    # Duplikate bei B entfernen
    seen, uniq = set(), []
    for t in treffer:
        sig = (t["kw"], t["seite"], t["auszug"][:60])
        if sig in seen:
            continue
        seen.add(sig); uniq.append(t)
    return bau, uniq, excluded


HEAD = ["Datum", "Ausgabe-Nr.", "Kategorie", "Relevanz", "Rubrik", "Gemeinde",
        "Bauherrschaft", "Bauvorhaben", "Bauplatz / Fundstelle",
        "Bemerkungen / Auszug", "Seite", "PDF-Link"]
WID = [11, 10, 20, 16, 22, 14, 30, 42, 34, 46, 6, 30]


def write_xlsx(bau, treffer, excluded, out, nr, datum, url):
    wb = Workbook(); ws = wb.active; ws.title = "Treffer"
    teal = PatternFill("solid", fgColor="1F6F6F")
    gray = PatternFill("solid", fgColor="EEF3F3")
    blue = PatternFill("solid", fgColor="D6E6F5")
    green = PatternFill("solid", fgColor="C9E7CE")
    ws.merge_cells("A1:L1")
    ws["A1"] = "Amtsblatt Uri – Baugesuche & Abwasser/Kläranlagen"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A2:L2")
    ws["A2"] = (f"Ausgabe Nr. {nr} · {datum} · Baugesuche: {len(bau)} · "
                f"Abwasser-Treffer: {len(treffer)} · ausgeschlossen (Solar/WP): {len(excluded)}")
    ws["A2"].font = Font(italic=True, color="555555")
    for c, h in enumerate(HEAD, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = teal
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = WID[c - 1]
    r = 5
    for e in bau:
        rel = f"Abwasser-Bezug ({e['kw']})" if e.get("kw") else "Baugesuch – prüfen"
        row = [datum, nr, "Baugesuch", rel, e.get("rubrik", ""), e.get("gemeinde", ""),
               e.get("bauherrschaft", ""), e.get("bauvorhaben", ""), e.get("bauplatz", ""),
               e.get("bemerkungen", ""), e.get("seite", ""), url]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=(c in (7, 8, 9, 10)), vertical="top")
            if r % 2 == 0: cell.fill = gray
        ws.cell(r, 4).fill = green if e.get("kw") else blue
        r += 1
    for t in treffer:
        row = [datum, nr, "Abwasser/Kläranlage", f"direkt ({t['kw']})", t.get("rubrik", ""),
               t.get("gemeinde", ""), "", "", "", t.get("auszug", ""), t.get("seite", ""), url]
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=(c == 10), vertical="top")
            if r % 2 == 0: cell.fill = gray
        ws.cell(r, 4).fill = green
        r += 1
    if not bau and not treffer:
        ws.merge_cells("A5:L5")
        ws["A5"] = "Keine relevanten Einträge in dieser Ausgabe."
        ws["A5"].font = Font(italic=True)
        r = 6
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:L{max(5, r - 1)}"

    if excluded:                      # Transparenz: was wurde weggefiltert
        ws2 = wb.create_sheet("Ausgeschlossen (Solar-WP)")
        ws2.append(["Gemeinde", "Bauherrschaft", "Bauvorhaben", "Seite"])
        for c in range(1, 5):
            ws2.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws2.cell(1, c).fill = teal
        for w, col in zip((14, 30, 46, 6), "ABCD"):
            ws2.column_dimensions[col].width = w
        for e in excluded:
            ws2.append([e.get("gemeinde", ""), e.get("bauherrschaft", ""),
                        e.get("bauvorhaben", ""), e.get("seite", "")])
    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("inp"); ap.add_argument("out")
    ap.add_argument("--nr", default="?"); ap.add_argument("--datum", default="?")
    ap.add_argument("--url", default="")
    a = ap.parse_args()
    if a.inp.lower().endswith(".json"):
        # Ergebnisse kommen aus dem Browser-Scan (browser_scan.js) — der
        # Rueckgabekanal des Browsers ist zu klein fuer den Volltext.
        import json
        d = json.load(open(a.inp, encoding="utf-8"))
        a.nr = d.get("nr", a.nr); a.datum = d.get("datum", a.datum)
        a.url = d.get("url", a.url)
        bau = [{"gemeinde": e.get("g", ""), "seite": e.get("s", ""),
                "rubrik": e.get("r", "Bauplanauflagen"),
                "bauherrschaft": e.get("h", ""), "bauvorhaben": e.get("v", ""),
                "bauplatz": e.get("p", ""), "bemerkungen": e.get("b", ""),
                "kw": e.get("k") or None} for e in d.get("bau", [])]
        treffer = [{"rubrik": e.get("r", "—"), "gemeinde": e.get("g", ""),
                    "seite": e.get("s", ""), "kw": e.get("k", ""),
                    "auszug": e.get("a", "")} for e in d.get("tre", [])]
        excluded = [{"gemeinde": e.get("g", ""), "bauherrschaft": e.get("h", ""),
                     "bauvorhaben": e.get("v", ""), "seite": e.get("s", "")}
                    for e in d.get("exc", [])]
    else:
        text = open(a.inp, encoding="utf-8").read()
        bau, treffer, excluded = scan(text)
    write_xlsx(bau, treffer, excluded, a.out, a.nr, a.datum, a.url)
    mit = sum(1 for e in bau if e.get("kw"))
    print(f"Baugesuche: {len(bau)} (davon mit Abwasser-Bezug: {mit})")
    print(f"Abwasser-/Kläranlagen-Treffer sonst: {len(treffer)}")
    print(f"Ausgeschlossen (reine Solar/Wärmepumpe): {len(excluded)}")
    for e in bau:
        flag = "!" if e.get("kw") else " "
        print(f" {flag} S.{e.get('seite','?'):>3} {e.get('gemeinde','')[:12]:12} {e.get('bauvorhaben','')[:60]}")
    for t in treffer:
        print(f" * S.{t.get('seite','?'):>3} {t.get('rubrik','')[:22]:22} <{t['kw']}>")
    for e in excluded:
        print(f" - S.{e.get('seite','?'):>3} {e.get('gemeinde','')[:12]:12} AUSGESCHLOSSEN: {e.get('bauvorhaben','')[:50]}")


if __name__ == "__main__":
    main()
